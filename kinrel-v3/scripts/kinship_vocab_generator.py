"""
Daxelo-Kinrel Deterministic Kinship Engine v3.0
Vocabulary Mapper Table Generator
=================================
Generates the full 5,396+ row kinship vocabulary database by multiplying a
seed table of distinct KinshipSignature-anchored English concepts by a
language/region matrix (18 languages).

Output:
  - /home/z/my-project/download/daxelo_kinrel_vocabulary.xlsx
  - /home/z/my-project/download/daxelo_kinrel_vocabulary.json
  - /home/z/my-project/download/daxelo_kinrel_vocabulary.csv

Each row carries the FULL runtime-only KinshipSignature so the engine can
look up the localized term without ever knowing what the term is.
"""

from __future__ import annotations
import csv
import json
import itertools
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl is required: pip install openpyxl")


# ---------------------------------------------------------------------------
# 1. KINSHIP SIGNATURE (runtime-only, never stored)
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class KinshipSignature:
    generationDelta: int            # -8 .. +8
    pathPattern: str                # e.g. "UP_PARENT_UP_PARENT"
    side: str                       # paternal | maternal | none
    consanguinity: str              # blood | half | step | adoptive | inLaw | foster | spiritual
    genderAnchor: str               # male | female | neutral
    seniority: str                  # elder | younger | twin | none
    removal: int                    # 0 .. 8
    doubleKinship: bool             # double first cousin etc.
    temporal: str = "current"       # current | former | late  (NEW — first-class temporal state)

    def signature_key(self) -> str:
        return "|".join([
            f"g={self.generationDelta}",
            f"p={self.pathPattern}",
            f"s={self.side}",
            f"c={self.consanguinity}",
            f"x={self.genderAnchor}",
            f"r={self.removal}",
            f"sn={self.seniority}",
            f"d={int(self.doubleKinship)}",
            f"t={self.temporal}",
        ])


# ---------------------------------------------------------------------------
# 2. SEED CONCEPTS — distinct English kinship terms + their signatures
# ---------------------------------------------------------------------------
# Each seed tuple: (canonicalId, englishTerm, signature, category, notes, variant_type)
# variant_type: "primary" (rank 0) | "regional" (rank 1) | "dialectal" (rank 2)

def _build_seed_concepts() -> List[Tuple[str, str, KinshipSignature, str, str, str]]:
    seeds: List[Tuple[str, str, KinshipSignature, str, str, str]] = []

    def add(canonical_id, term, sig, category, notes="", variant_type="primary"):
        seeds.append((canonical_id, term, sig, category, notes, variant_type))

    # ---- 2.1 Direct ancestors (UP_PARENT chain), depth 1..8 ----
    # English convention:
    #   depth 1 → father / mother
    #   depth 2 → grandfather / grandmother        (NO "great-" prefix)
    #   depth 3 → great-grandfather / great-grandmother
    #   depth 4 → 2nd great-grandfather
    #   depth 5 → 3rd great-grandfather
    #   ...
    ancestor_ordinals = [
        (1, ""),                (2, "grand"),
        (3, "great-grand"),     (4, "2nd great-grand"),
        (5, "3rd great-grand"), (6, "4th great-grand"),
        (7, "5th great-grand"), (8, "6th great-grand"),
    ]
    for depth, prefix in ancestor_ordinals:
        pattern = "_".join(["UP_PARENT"] * depth)
        gen = -depth
        for side, sname in [("paternal", "paternal"), ("maternal", "maternal")]:
            for gender, gname in [("male", "father"), ("female", "mother")]:
                term = f"{prefix}{gname}"
                add("DERIVED", term,
                    KinshipSignature(gen, pattern, side, "blood", gender, "none", 0, False),
                    "direct_ancestor", f"{depth}-up on {side} side")

    # ---- 2.2 Direct descendants (DOWN_CHILD chain), depth 1..8 ----
    # English convention:
    #   depth 1 → son / daughter
    #   depth 2 → grandson / granddaughter
    #   depth 3 → great-grandson / great-granddaughter
    #   ...
    desc_ordinals = [
        (1, ""),                (2, "grand"),
        (3, "great-grand"),     (4, "2nd great-grand"),
        (5, "3rd great-grand"), (6, "4th great-grand"),
        (7, "5th great-grand"), (8, "6th great-grand"),
    ]
    for depth, prefix in desc_ordinals:
        pattern = "_".join(["DOWN_CHILD"] * depth)
        gen = +depth
        for gender, gname in [("male", "son"), ("female", "daughter")]:
            term = f"{prefix}{gname}"
            add("DERIVED", term,
                KinshipSignature(gen, pattern, "none", "blood", gender, "none", 0, False),
                "direct_descendant", f"{depth}-down")

    # ---- 2.3 Siblings (UP_PARENT_DOWN_CHILD), all consanguinity & seniority ----
    sib_consanguinity = [
        ("blood", "Full",  ["elder", "younger", "twin", "none"]),
        ("half",  "Half",  ["elder", "younger", "none"]),
        ("step",  "Step",  ["none"]),
        ("adoptive", "Adoptive", ["none"]),
        ("foster",   "Foster",   ["none"]),
    ]
    for consang, label, seniorities in sib_consanguinity:
        for gender, gname in [("male", "brother"), ("female", "sister")]:
            for sen in seniorities:
                if sen == "none":
                    term = f"{label} {gname}".replace("Full ", "").strip().title()
                else:
                    term = f"{sen.capitalize()} {label} {gname}".replace("Full ", "").strip().title()
                add("DERIVED", term,
                    KinshipSignature(0, "UP_PARENT_DOWN_CHILD", "none", consang, gender, sen, 0, False),
                    "sibling", f"{consang} sibling, {sen}")

    # ---- 2.4 Aunts / Uncles (UP_PARENT UP_PARENT DOWN_CHILD), depth 2..8 ----
    # English convention:
    #   depth 2 UP → uncle / aunt                          (1 generation above parent)
    #   depth 3 UP → great-uncle / great-aunt              (grandparent's sibling)
    #   depth 4 UP → 2nd great-uncle
    #   ...
    # Gen -1 (uncle), -2 (great-uncle), ..., -7 (6th great-uncle)
    uncle_ordinals = [
        ("", 2),
        ("great-", 3),
        ("2nd great-", 4),
        ("3rd great-", 5),
        ("4th great-", 6),
        ("5th great-", 7),
        ("6th great-", 8),
    ]
    for prefix, depth in uncle_ordinals:
        pattern = "_".join(["UP_PARENT"] * depth) + "_DOWN_CHILD"
        gen = -(depth - 1)
        for side in ["paternal", "maternal"]:
            for gender, gname in [("male", "uncle"), ("female", "aunt")]:
                term = f"{prefix}{gname}".strip("-")
                # Special: in some English dialects great-uncle = granduncle, both kept
                add("DERIVED", term,
                    KinshipSignature(gen, pattern, side, "blood", gender, "none", 0, False),
                    "aunt_uncle", f"{depth-1}-up collateral, {side}")

    # Alternate regional forms (granduncle / grandaunt, great-granduncle...)
    grand_alt_ordinals = [
        ("grand", 2), ("great-grand", 3), ("2nd great-grand", 4),
        ("3rd great-grand", 5), ("4th great-grand", 6),
        ("5th great-grand", 7), ("6th great-grand", 8),
    ]
    for alt_prefix, depth in grand_alt_ordinals:
        pattern = "_".join(["UP_PARENT"] * depth) + "_DOWN_CHILD"
        gen = -(depth - 1)
        for side in ["paternal", "maternal"]:
            for gender, gname in [("male", "uncle"), ("female", "aunt")]:
                term = f"{alt_prefix}{gname}"
                add("DERIVED", term,
                    KinshipSignature(gen, pattern, side, "blood", gender, "none", 0, False),
                    "aunt_uncle", f"{depth-1}-up collateral alt form", variant_type="regional")

    # ---- 2.5 Nieces / Nephews (UP_PARENT DOWN_CHILD DOWN_CHILD) ----
    niece_ordinals = [
        ("", 2), ("great-", 3), ("2nd great-", 4), ("3rd great-", 5),
        ("4th great-", 6), ("5th great-", 7), ("6th great-", 8),
    ]
    for prefix, depth in niece_ordinals:
        pattern = "UP_PARENT" + "_DOWN_CHILD" * depth
        gen = +(depth - 1)
        for gender, gname in [("male", "nephew"), ("female", "niece")]:
            term = f"{prefix}{gname}".strip("-")
            add("DERIVED", term,
                KinshipSignature(gen, pattern, "none", "blood", gender, "none", 0, False),
                "niece_nephew", f"{depth-1}-down collateral")

    # ---- 2.6 Cousins (UP^N DOWN^N, with removals) ----
    # 1st cousin: UP^2 DOWN^2 (gen 0, removal 0)
    # 1st cousin once removed: UP^2 DOWN^3 (gen +1, removal 1)
    # 2nd cousin: UP^3 DOWN^3 (gen 0, removal 0)
    # ...
    for cousin_order in range(1, 9):           # 1st .. 8th cousin
        up_depth = 1 + cousin_order              # 2,3,4,5,6,7,8,9 -> capped at 8
        if up_depth > 8:
            continue
        for removal in range(0, 9):              # 0 .. 8 times removed
            down_depth = up_depth + removal
            total_depth = up_depth + down_depth
            if total_depth > 8:
                continue
            pattern = "_".join(["UP_PARENT"] * up_depth) + "__" + "_".join(["DOWN_CHILD"] * down_depth)
            gen = removal
            ordinal_word = _ordinal(cousin_order)
            for side in ["paternal", "maternal"]:
                for gender, gname in [("male", "cousin (m)"), ("female", "cousin (f)"), ("neutral", "cousin")]:
                    base = f"{ordinal_word} cousin"
                    if removal > 0:
                        base += f" {_times_removed(removal)}"
                    if gender != "neutral":
                        base += f" [{gname}]"
                    add("DERIVED", base,
                        KinshipSignature(gen, pattern, side, "blood", gender, "none", removal, False),
                        "cousin", f"order={cousin_order}, removal={removal}, {side}")
            # Double cousin variants (only same order, removal 0, blood)
            if removal == 0:
                for gender, gname in [("male", "cousin (m)"), ("female", "cousin (f)"), ("neutral", "cousin")]:
                    base = f"double {ordinal_word} cousin"
                    if gender != "neutral":
                        base += f" [{gname}]"
                    add("DERIVED", base,
                        KinshipSignature(gen, pattern, "paternal", "blood", gender, "none", 0, True),
                        "cousin", f"double {ordinal_word} cousin")

    # ---- 2.7 Spouse & in-laws (SPOUSE_* / *_SPOUSE) ----
    # Spouse
    for gender, gname in [("male", "husband"), ("female", "wife"), ("neutral", "spouse")]:
        add("SPOUSE", gname,
            KinshipSignature(0, "SPOUSE", "none", "inLaw", gender, "none", 0, False),
            "spouse", "direct spouse")

    # In-laws by spouse's family
    # Father/mother-in-law: SPOUSE_UP_PARENT
    for gender, gname in [("male", "father-in-law"), ("female", "mother-in-law")]:
        add("DERIVED", gname,
            KinshipSignature(-1, "SPOUSE_UP_PARENT", "none", "inLaw", gender, "none", 0, False),
            "in_law", "spouse's parent")

    # Son/daughter-in-law: SPOUSE of own child = UP_PARENT inverse then SPOUSE -> use UP_PARENT_DOWN_CHILD_SPOUSE pattern
    # We use: UP_PARENT_DOWN_CHILD_SPOUSE? Actually it's: child is via DOWN_CHILD, then their spouse. Pattern = DOWN_CHILD_SPOUSE
    for gender, gname in [("male", "son-in-law"), ("female", "daughter-in-law")]:
        add("DERIVED", gname,
            KinshipSignature(+1, "DOWN_CHILD_SPOUSE", "none", "inLaw", gender, "none", 0, False),
            "in_law", "child's spouse")

    # Brother/sister-in-law (spouse's sibling): SPOUSE_UP_PARENT_DOWN_CHILD
    for sen in ["elder", "younger", "none"]:
        for gender, gname in [("male", "brother-in-law"), ("female", "sister-in-law")]:
            term = gname if sen == "none" else f"{sen} {gname}"
            add("DERIVED", term,
                KinshipSignature(0, "SPOUSE_UP_PARENT_DOWN_CHILD", "none", "inLaw", gender, sen, 0, False),
                "in_law", f"spouse's sibling ({sen})")

    # Sibling's spouse (own sibling's spouse): UP_PARENT_DOWN_CHILD_SPOUSE
    for gender, gname in [("male", "brother-in-law"), ("female", "sister-in-law")]:
        add("DERIVED", f"sibling's {gname.replace('-in-law','')}-in-law",
            KinshipSignature(0, "UP_PARENT_DOWN_CHILD_SPOUSE", "none", "inLaw", gender, "none", 0, False),
            "in_law", "own sibling's spouse")

    # Grandparent-in-law: SPOUSE_UP_PARENT_UP_PARENT
    for depth, prefix in ancestor_ordinals[1:]:
        pattern = "SPOUSE_" + "_".join(["UP_PARENT"] * depth)
        gen = -depth
        for gender, gname in [("male", "grandfather-in-law"), ("female", "grandmother-in-law")]:
            term = f"{prefix}{gname}".strip("-")
            add("DERIVED", term,
                KinshipSignature(gen, pattern, "none", "inLaw", gender, "none", 0, False),
                "in_law", f"spouse's {prefix}grandparent".strip("-"))

    # Grandchild-in-law: DOWN_CHILD_DOWN_CHILD_SPOUSE
    for depth, prefix in desc_ordinals[1:]:
        pattern = "_".join(["DOWN_CHILD"] * depth) + "_SPOUSE"
        gen = +depth
        for gender, gname in [("male", "grandson-in-law"), ("female", "granddaughter-in-law")]:
            term = f"{prefix}{gname}".strip("-")
            add("DERIVED", term,
                KinshipSignature(gen, pattern, "none", "inLaw", gender, "none", 0, False),
                "in_law", f"grandchild's spouse")

    # Aunt/uncle-in-law (spouse's aunt/uncle): SPOUSE_UP_PARENT_UP_PARENT_DOWN_CHILD
    for prefix, depth in uncle_ordinals:
        pattern = "SPOUSE_" + "_".join(["UP_PARENT"] * depth) + "_DOWN_CHILD"
        gen = -(depth - 1)
        for gender, gname in [("male", "uncle-in-law"), ("female", "aunt-in-law")]:
            term = f"{prefix}{gname}".strip("-")
            add("DERIVED", term,
                KinshipSignature(gen, pattern, "none", "inLaw", gender, "none", 0, False),
                "in_law", "spouse's aunt/uncle")

    # Niece/nephew-in-law (spouse's niece/nephew): SPOUSE_UP_PARENT_DOWN_CHILD_DOWN_CHILD
    for prefix, depth in niece_ordinals:
        pattern = "SPOUSE_UP_PARENT" + "_DOWN_CHILD" * depth
        gen = +(depth - 1)
        for gender, gname in [("male", "nephew-in-law"), ("female", "niece-in-law")]:
            term = f"{prefix}{gname}".strip("-")
            add("DERIVED", term,
                KinshipSignature(gen, pattern, "none", "inLaw", gender, "none", 0, False),
                "in_law", "spouse's niece/nephew")

    # Cousin-in-law (spouse's cousin): SPOUSE_UP^N_DOWN^N
    for cousin_order in range(1, 7):
        up_depth = 1 + cousin_order
        for removal in range(0, 4):
            down_depth = up_depth + removal
            if up_depth + down_depth > 8:
                continue
            pattern = "SPOUSE_" + "_".join(["UP_PARENT"] * up_depth) + "__" + "_".join(["DOWN_CHILD"] * down_depth)
            gen = removal
            for gender, gname in [("male", "cousin-in-law (m)"), ("female", "cousin-in-law (f)"), ("neutral", "cousin-in-law")]:
                base = f"{_ordinal(cousin_order)} cousin-in-law"
                if removal > 0:
                    base += f" {_times_removed(removal)}"
                if gender != "neutral":
                    base += f" [{gname.split('(')[1].strip(') ')}]"
                add("DERIVED", base,
                    KinshipSignature(gen, pattern, "none", "inLaw", gender, "none", removal, False),
                    "in_law", "spouse's cousin")

    # ---- 2.8 Step-family ----
    step_pairs = [
        ("stepfather", -1, "UP_STEP_PARENT", "male"),
        ("stepmother", -1, "UP_STEP_PARENT", "female"),
        ("stepson", +1, "DOWN_STEP_CHILD", "male"),
        ("stepdaughter", +1, "DOWN_STEP_CHILD", "female"),
        ("stepbrother", 0, "UP_STEP_PARENT_DOWN_STEP_CHILD", "male"),
        ("stepsister", 0, "UP_STEP_PARENT_DOWN_STEP_CHILD", "female"),
    ]
    for term, gen, pattern, gender in step_pairs:
        add("STEP_PARENT" if "STEP_PARENT" in pattern and gen < 0 else "DERIVED",
            term,
            KinshipSignature(gen, pattern, "none", "step", gender, "none", 0, False),
            "step_family", "step relationship")

    # Step-grandparents / step-grandchildren
    for depth, prefix in ancestor_ordinals[1:]:
        pattern = "UP_STEP_PARENT_" + "_".join(["UP_PARENT"] * (depth - 1)) if depth > 1 else "UP_STEP_PARENT"
        # Simplified — treat as step variant of grandparent
        pattern = "UP_STEP_PARENT_" + "_".join(["UP_PARENT"] * (depth - 1))
        gen = -depth
        for gender, gname in [("male", "step-grandfather"), ("female", "step-grandmother")]:
            term = f"{prefix}{gname}".strip("-")
            add("DERIVED", term,
                KinshipSignature(gen, pattern, "none", "step", gender, "none", 0, False),
                "step_family", "step grandparent")

    for depth, prefix in desc_ordinals[1:]:
        pattern = "_".join(["DOWN_CHILD"] * (depth - 1)) + "_DOWN_STEP_CHILD"
        gen = +depth
        for gender, gname in [("male", "step-grandson"), ("female", "step-granddaughter")]:
            term = f"{prefix}{gname}".strip("-")
            add("DERIVED", term,
                KinshipSignature(gen, pattern, "none", "step", gender, "none", 0, False),
                "step_family", "step grandchild")

    # Step-aunt/uncle, step-niece/nephew
    for gender, gname in [("male", "step-uncle"), ("female", "step-aunt")]:
        add("DERIVED", gname,
            KinshipSignature(-1, "UP_STEP_PARENT_UP_PARENT_DOWN_CHILD", "none", "step", gender, "none", 0, False),
            "step_family", "step-uncle/aunt")
    for gender, gname in [("male", "step-nephew"), ("female", "step-niece")]:
        add("DERIVED", gname,
            KinshipSignature(+1, "UP_PARENT_DOWN_CHILD_DOWN_STEP_CHILD", "none", "step", gender, "none", 0, False),
            "step_family", "step-nephew/niece")

    # ---- 2.9 Adoptive family ----
    adoptive_pairs = [
        ("adoptive father", -1, "UP_ADOPTIVE_PARENT", "male"),
        ("adoptive mother", -1, "UP_ADOPTIVE_PARENT", "female"),
        ("adoptive son", +1, "DOWN_ADOPTIVE_CHILD", "male"),
        ("adoptive daughter", +1, "DOWN_ADOPTIVE_CHILD", "female"),
        ("adoptive brother", 0, "UP_ADOPTIVE_PARENT_DOWN_ADOPTIVE_CHILD", "male"),
        ("adoptive sister", 0, "UP_ADOPTIVE_PARENT_DOWN_ADOPTIVE_CHILD", "female"),
    ]
    for term, gen, pattern, gender in adoptive_pairs:
        add("ADOPTIVE_PARENT" if "ADOPTIVE_PARENT" in pattern and gen < 0 else "DERIVED",
            term,
            KinshipSignature(gen, pattern, "none", "adoptive", gender, "none", 0, False),
            "adoptive_family", "adoptive relationship")

    # Adoptive grandparents / grandchildren
    for depth, prefix in ancestor_ordinals[1:]:
        pattern = "UP_ADOPTIVE_PARENT_" + "_".join(["UP_PARENT"] * (depth - 1))
        gen = -depth
        for gender, gname in [("male", "adoptive grandfather"), ("female", "adoptive grandmother")]:
            term = f"{prefix}{gname}".strip("-")
            add("DERIVED", term,
                KinshipSignature(gen, pattern, "none", "adoptive", gender, "none", 0, False),
                "adoptive_family", "adoptive grandparent")
    for depth, prefix in desc_ordinals[1:]:
        pattern = "_".join(["DOWN_CHILD"] * (depth - 1)) + "_DOWN_ADOPTIVE_CHILD"
        gen = +depth
        for gender, gname in [("male", "adoptive grandson"), ("female", "adoptive granddaughter")]:
            term = f"{prefix}{gname}".strip("-")
            add("DERIVED", term,
                KinshipSignature(gen, pattern, "none", "adoptive", gender, "none", 0, False),
                "adoptive_family", "adoptive grandchild")

    # ---- 2.10 Foster family ----
    foster_pairs = [
        ("foster father", -1, "UP_FOSTER_PARENT", "male"),
        ("foster mother", -1, "UP_FOSTER_PARENT", "female"),
        ("foster son", +1, "DOWN_FOSTER_CHILD", "male"),
        ("foster daughter", +1, "DOWN_FOSTER_CHILD", "female"),
        ("foster brother", 0, "UP_FOSTER_PARENT_DOWN_FOSTER_CHILD", "male"),
        ("foster sister", 0, "UP_FOSTER_PARENT_DOWN_FOSTER_CHILD", "female"),
    ]
    for term, gen, pattern, gender in foster_pairs:
        add("DERIVED", term,
            KinshipSignature(gen, pattern, "none", "foster", gender, "none", 0, False),
            "foster_family", "foster relationship")

    # ---- 2.11 Spiritual / god-family ----
    spiritual_pairs = [
        ("godfather", -1, "UP_SPIRITUAL_PARENT", "male"),
        ("godmother", -1, "UP_SPIRITUAL_PARENT", "female"),
        ("godson", +1, "DOWN_SPIRITUAL_CHILD", "male"),
        ("goddaughter", +1, "DOWN_SPIRITUAL_CHILD", "female"),
        ("godbrother", 0, "UP_SPIRITUAL_PARENT_DOWN_SPIRITUAL_CHILD", "male"),
        ("godsister", 0, "UP_SPIRITUAL_PARENT_DOWN_SPIRITUAL_CHILD", "female"),
    ]
    for term, gen, pattern, gender in spiritual_pairs:
        add("DERIVED", term,
            KinshipSignature(gen, pattern, "none", "spiritual", gender, "none", 0, False),
            "spiritual_family", "god-relationship")

    # ---- 2.12 Ex-family ( divorced / former ) ----
    # Ex-family now uses first-class temporal="former" field in the signature,
    # so they have DISTINCT signature_keys from the current relationship.
    # The engine can deterministically resolve "ex-husband" without any
    # variant_rank fallback. The temporal flag is read from the stored edge.
    ex_pairs = [
        ("ex-husband", 0, "SPOUSE", "male"),
        ("ex-wife", 0, "SPOUSE", "female"),
        ("ex-father-in-law", -1, "SPOUSE_UP_PARENT", "male"),
        ("ex-mother-in-law", -1, "SPOUSE_UP_PARENT", "female"),
        ("ex-brother-in-law", 0, "SPOUSE_UP_PARENT_DOWN_CHILD", "male"),
        ("ex-sister-in-law", 0, "SPOUSE_UP_PARENT_DOWN_CHILD", "female"),
        ("ex-son-in-law", +1, "DOWN_CHILD_SPOUSE", "male"),
        ("ex-daughter-in-law", +1, "DOWN_CHILD_SPOUSE", "female"),
    ]
    for term, gen, pattern, gender in ex_pairs:
        add("DERIVED", term,
            KinshipSignature(gen, pattern, "none", "inLaw", gender, "none", 0, False, temporal="former"),
            "ex_family", "former relationship — temporal=former in signature",
            variant_type="primary")

    # ---- 2.13 Late (deceased) family — first-class temporal="late" ----
    # Used for deceased parents, grandparents, spouse, siblings.
    # The path pattern is identical to the living variant, but temporal="late"
    # produces a distinct signature_key so the engine can resolve terms like
    # "late father" / "deceased husband" deterministically.
    late_pairs = [
        ("late father", -1, "UP_PARENT", "male", "paternal"),
        ("late mother", -1, "UP_PARENT", "female", "maternal"),
        ("late husband", 0, "SPOUSE", "male", "none"),
        ("late wife", 0, "SPOUSE", "female", "none"),
        ("late grandfather", -2, "UP_PARENT_UP_PARENT", "male", "paternal"),
        ("late grandmother", -2, "UP_PARENT_UP_PARENT", "female", "paternal"),
        ("late brother", 0, "UP_PARENT_DOWN_CHILD", "male", "none"),
        ("late sister", 0, "UP_PARENT_DOWN_CHILD", "female", "none"),
        ("late son", +1, "DOWN_CHILD", "male", "none"),
        ("late daughter", +1, "DOWN_CHILD", "female", "none"),
    ]
    for term, gen, pattern, gender, side in late_pairs:
        add("DERIVED", term,
            KinshipSignature(gen, pattern, side, "blood", gender, "none", 0, False, temporal="late"),
            "late_family", "deceased relative — temporal=late in signature",
            variant_type="primary")

    return seeds


def _ordinal(n: int) -> str:
    words = ["zeroth", "first", "second", "third", "fourth", "fifth",
             "sixth", "seventh", "eighth", "ninth", "tenth"]
    if 0 <= n < len(words):
        return words[n]
    return f"{n}th"


def _times_removed(n: int) -> str:
    if n == 1:
        return "once removed"
    if n == 2:
        return "twice removed"
    if n == 3:
        return "thrice removed"
    return f"{_ordinal(n)} removed"


# ---------------------------------------------------------------------------
# 3. LANGUAGE DICTIONARIES
# ---------------------------------------------------------------------------
# Each language provides root words + composition rules.
# Roots: father, mother, parent, son, daughter, child, husband, wife, spouse,
#        brother, sister, sibling, grandfather, grandmother, grandson, granddaughter,
#        uncle, aunt, nephew, niece, cousin, self
# Modifiers: great, half, step, adoptive, foster, in_law, elder, younger, twin,
#            paternal, maternal, double, removed
#
# Composition: term = compose(language_dict, english_term, signature)
# Uses simple rule-based translation. Falls back to English root + language modifier.

LANGUAGES: Dict[str, Dict[str, str]] = {
    "en": {  # English (reference)
        "name": "English",
        "father": "father", "mother": "mother", "parent": "parent",
        "son": "son", "daughter": "daughter", "child": "child",
        "husband": "husband", "wife": "wife", "spouse": "spouse",
        "brother": "brother", "sister": "sister", "sibling": "sibling",
        "grandfather": "grandfather", "grandmother": "grandmother",
        "grandson": "grandson", "granddaughter": "granddaughter",
        "uncle": "uncle", "aunt": "aunt",
        "nephew": "nephew", "niece": "niece",
        "cousin": "cousin",
        "great": "great-", "half": "half-", "step": "step-",
        "adoptive": "adoptive ", "foster": "foster ",
        "in_law": "-in-law", "elder": "elder ", "younger": "younger ",
        "twin": "twin ", "paternal": "paternal ", "maternal": "maternal ",
        "double": "double ", "ex": "ex-",
        "god": "god",
    },
    "hi": {  # Hindi (Devanagari) — kinship-rich
        "name": "Hindi",
        "father": "पिता", "mother": "माता", "parent": "अभिभावक",
        "son": "पुत्र", "daughter": "पुत्री", "child": "संतान",
        "husband": "पति", "wife": "पत्नी", "spouse": "जीवनसाथी",
        "brother": "भाई", "sister": "बहन", "sibling": "सहोदर",
        "grandfather": "दादा", "grandmother": "दादी",
        "grandson": "पोता", "granddaughter": "पोती",
        "uncle": "चाचा", "aunt": "बुआ",
        "nephew": "भतीजा", "niece": "भतीजी",
        "cousin": "चचेरा भाई",
        "great": "पर-", "half": "सौतेला ", "step": "सौतेला ",
        "adoptive": "दत्तक ", "foster": "पालक ",
        "in_law": " सुलभ", "elder": "बड़ा ", "younger": "छोटा ",
        "twin": "जुड़वा ", "paternal": "पैतृक ", "maternal": "मातृ ",
        "double": "द्विगुण ", "ex": "पूर्व-",
        "god": "धर्म-",
    },
    "ta": {  # Tamil
        "name": "Tamil",
        "father": "தந்தை", "mother": "தாய்", "parent": "பெற்றோர்",
        "son": "மகன்", "daughter": "மகள்", "child": "குழந்தை",
        "husband": "கணவர்", "wife": "மனைவி", "spouse": "வாழ்க்கைத் துணை",
        "brother": "சகோதரன்", "sister": "சகோதரி", "sibling": "சகோதரர்",
        "grandfather": "தாத்தா", "grandmother": "பாட்டி",
        "grandson": "பேரப்பிள்ளை", "granddaughter": "பேரப்பெண்",
        "uncle": "மாமா", "aunt": "அத்தை",
        "nephew": "மருமகன்", "niece": "மருமகள்",
        "cousin": "உறவினர்",
        "great": "பெரிய-", "half": "அரை-", "step": "வளர்ப்பு-",
        "adoptive": "தத்தெடுப்பு ", "foster": "பாதுகாப்பு ",
        "in_law": " மருமகன்", "elder": "மூத்த ", "younger": "இளைய ",
        "twin": "இரட்டை ", "paternal": "தந்தை வழி ", "maternal": "தாய் வழி ",
        "double": "இரட்டை ", "ex": "முன்னாள்-",
        "god": "கடவுள்-",
    },
    "te": {  # Telugu
        "name": "Telugu",
        "father": "తండ్రి", "mother": "తల్లి", "parent": "తల్లిదండ్రులు",
        "son": "కుమారుడు", "daughter": "కుమార్తె", "child": "బిడ్డ",
        "husband": "భర్త", "wife": "భార్య", "spouse": "జీవిత భాగస్వామి",
        "brother": "సోదరుడు", "sister": "సోదరి", "sibling": "సోదరులు",
        "grandfather": "తాత", "grandmother": "అమ్మమ్మ",
        "grandson": "మనవడు", "granddaughter": "మనవరాలు",
        "uncle": "మామయ్య", "aunt": "పిన్ని",
        "nephew": "మేనల్లుడు", "niece": "మేనకోడలు",
        "cousin": "బంధువు",
        "great": "గొప్ప-", "half": "సగం-", "step": "సవతి-",
        "adoptive": "దత్తత ", "foster": "సంరక్షణ ",
        "in_law": " అల్లుడు", "elder": "పెద్ద ", "younger": "చిన్న ",
        "twin": "కవల ", "paternal": "తండ్రి వారసత్వ ", "maternal": "తల్లి వారసత్వ ",
        "double": "రెట్టింపు ", "ex": "మాజీ-",
        "god": "దేవుని-",
    },
    "bn": {  # Bengali
        "name": "Bengali",
        "father": "পিতা", "mother": "মাতা", "parent": "অভিভাবক",
        "son": "পুত্র", "daughter": "কন্যা", "child": "সন্তান",
        "husband": "স্বামী", "wife": "স্ত্রী", "spouse": "জীবনসঙ্গী",
        "brother": "ভাই", "sister": "বোন", "sibling": "ভাইবোন",
        "grandfather": "দাদু", "grandmother": "দিদা",
        "grandson": "নাতি", "granddaughter": "নাতনি",
        "uncle": "কাকা", "aunt": "পিসি",
        "nephew": "ভাগ্নে", "niece": "ভাগ্নি",
        "cousin": "কাজিন",
        "great": "পর-", "half": "সৎ-", "step": "সৎ-",
        "adoptive": "দত্তক ", "foster": "পালক ",
        "in_law": " জামাই", "elder": "বড় ", "younger": "ছোট ",
        "twin": "যমজ ", "paternal": "পৈতৃক ", "maternal": "মাতৃ ",
        "double": "দ্বিগুণ ", "ex": "প্রাক্তন-",
        "god": "ধর্ম-",
    },
    "mr": {  # Marathi
        "name": "Marathi",
        "father": "वडील", "mother": "आई", "parent": "पालक",
        "son": "मुलगा", "daughter": "मुलगी", "child": "मूल",
        "husband": "पती", "wife": "पत्नी", "spouse": "जोडीदार",
        "brother": "भाऊ", "sister": "बहीण", "sibling": "भावंड",
        "grandfather": "आजोबा", "grandmother": "आजी",
        "grandson": "नातू", "granddaughter": "नात",
        "uncle": "काका", "aunt": "आत्या",
        "nephew": "भाचा", "niece": "भाची",
        "cousin": "चुलत भाऊ",
        "great": "पर-", "half": "सावत्र-", "step": "सावत्र-",
        "adoptive": "दत्तक ", "foster": "पालक ",
        "in_law": " जावई", "elder": "थोरले ", "younger": "धाकटे ",
        "twin": "जुळे ", "paternal": "पैतृक ", "maternal": "मातृ ",
        "double": "दुहेरी ", "ex": "माजी-",
        "god": "धर्म-",
    },
    "ml": {  # Malayalam
        "name": "Malayalam",
        "father": "അച്ഛൻ", "mother": "അമ്മ", "parent": "മാതാപിതാക്കൾ",
        "son": "മകൻ", "daughter": "മകൾ", "child": "കുട്ടി",
        "husband": "ഭർത്താവ്", "wife": "ഭാര്യ", "spouse": "ജീവിതപങ്കാളി",
        "brother": "സഹോദരൻ", "sister": "സഹോദരി", "sibling": "സഹോദരങ്ങൾ",
        "grandfather": "അപ്പൂപ്പൻ", "grandmother": "അമ്മൂമ്മ",
        "grandson": "പേരമകൻ", "granddaughter": "പേരമകൾ",
        "uncle": "അമ്മാവൻ", "aunt": "അമ്മായി",
        "nephew": "അനന്തരവൻ", "niece": "അനന്തരവൾ",
        "cousin": "കസിൻ",
        "great": "വലിയ-", "half": "പകുതി-", "step": "രണ്ടാം-",
        "adoptive": "ദത്തെടുത്ത ", "foster": "സംരക്ഷണ ",
        "in_law": " മരുമകൻ", "elder": "മൂത്ത ", "younger": "ഇളയ ",
        "twin": "ഇരട്ട ", "paternal": "പിതൃ ", "maternal": "മാതൃ ",
        "double": "ഇരട്ട ", "ex": "മുൻ-",
        "god": "ദൈവ-",
    },
    "kn": {  # Kannada
        "name": "Kannada",
        "father": "ತಂದೆ", "mother": "ತಾಯಿ", "parent": "ಪೋಷಕರು",
        "son": "ಮಗ", "daughter": "ಮಗಳು", "child": "ಮಗು",
        "husband": "ಗಂಡ", "wife": "ಹೆಂಡತಿ", "spouse": "ಜೀವನ ಸಂಗಾತಿ",
        "brother": "ಸಹೋದರ", "sister": "ಸಹೋದರಿ", "sibling": "ಒಡಹುಟ್ಟಿದವರು",
        "grandfather": "ಅಜ್ಜ", "grandmother": "ಅಜ್ಜಿ",
        "grandson": "ಮೊಮ್ಮಗ", "granddaughter": "ಮೊಮ್ಮಗಳು",
        "uncle": "ಮಾಮ", "aunt": "ಅತ್ತೆ",
        "nephew": "ಸೋದರ ಮಗ", "niece": "ಸೋದರ ಮಗಳು",
        "cousin": "ಸೋದರ ಸಂಬಂಧಿ",
        "great": "ಮಹಾ-", "half": "ಅರ್ಧ-", "step": "ಮಲ-",
        "adoptive": "ದತ್ತು ", "foster": "ಪಾಲಕ ",
        "in_law": " ಅಳಿಯ", "elder": "ಹಿರಿಯ ", "younger": "ಕಿರಿಯ ",
        "twin": "ಅವಳಿ ", "paternal": "ತಂದೆಯ ", "maternal": "ತಾಯಿಯ ",
        "double": "ಎರಡರಷ್ಟು ", "ex": "ಮಾಜಿ-",
        "god": "ಧರ್ಮ-",
    },
    "gu": {  # Gujarati
        "name": "Gujarati",
        "father": "પિતા", "mother": "માતા", "parent": "વાલી",
        "son": "દીકરો", "daughter": "દીકરી", "child": "બાળક",
        "husband": "પતિ", "wife": "પત્ની", "spouse": "જીવનસાથી",
        "brother": "ભાઈ", "sister": "બહેન", "sibling": "ભાઈ-બહેન",
        "grandfather": "દાદાજી", "grandmother": "દાદીમા",
        "grandson": "પૌત્ર", "granddaughter": "પૌત્રી",
        "uncle": "કાકા", "aunt": "કાકી",
        "nephew": "ભત્રીજો", "niece": "ભત્રીજી",
        "cousin": "કઝિન",
        "great": "પર-", "half": "સાવકા-", "step": "સાવકા-",
        "adoptive": "દત્તક ", "foster": "પાલક ",
        "in_law": " જમાઈ", "elder": "મોટા ", "younger": "નાના ",
        "twin": "જોડિયા ", "paternal": "પૈતૃક ", "maternal": "માતૃ ",
        "double": "દ્વિગુણ ", "ex": "ભૂતપૂર્વ-",
        "god": "ધર્મ-",
    },
    "pa": {  # Punjabi
        "name": "Punjabi",
        "father": "ਪਿਤਾ", "mother": "ਮਾਂ", "parent": "ਮਾਪੇ",
        "son": "ਪੁੱਤਰ", "daughter": "ਧੀ", "child": "ਬੱਚਾ",
        "husband": "ਪਤੀ", "wife": "ਪਤਨੀ", "spouse": "ਜੀਵਨ ਸਾਥੀ",
        "brother": "ਭਰਾ", "sister": "ਭੈਣ", "sibling": "ਭੈਣ-ਭਰਾ",
        "grandfather": "ਦਾਦਾ", "grandmother": "ਦਾਦੀ",
        "grandson": "ਪੋਤਾ", "granddaughter": "ਪੋਤੀ",
        "uncle": "ਚਾਚਾ", "aunt": "ਬੂਆ",
        "nephew": "ਭਤੀਜਾ", "niece": "ਭਤੀਜੀ",
        "cousin": "ਚਚੇਰਾ ਭਰਾ",
        "great": "ਪੜਦਾਦਾ-", "half": "ਮਤਰੇਯਾ-", "step": "ਮਤਰੇਯਾ-",
        "adoptive": "ਗੋਦ ", "foster": "ਪਾਲਣ ",
        "in_law": " ਜੁਆਈ", "elder": "ਵੱਡਾ ", "younger": "ਛੋਟਾ ",
        "twin": "ਜੁੜਵਾਂ ", "paternal": "ਪਿਤਾ ਪੱਖੀ ", "maternal": "ਮਾਂ ਪੱਖੀ ",
        "double": "ਦੋਹਰਾ ", "ex": "ਸਾਬਕਾ-",
        "god": "ਧਰਮ-",
    },
    "ur": {  # Urdu
        "name": "Urdu",
        "father": "والد", "mother": "والدہ", "parent": "والدین",
        "son": "بیٹا", "daughter": "بیٹی", "child": "بچہ",
        "husband": "شوہر", "wife": "بیوی", "spouse": "شریک حیات",
        "brother": "بھائی", "sister": "بہن", "sibling": "بھائی بہن",
        "grandfather": "دادا", "grandmother": "دادی",
        "grandson": "پوتا", "granddaughter": "پوتی",
        "uncle": "چچا", "aunt": "پھوپھی",
        "nephew": "بھتیجا", "niece": "بھتیجی",
        "cousin": "چچا زاد بھائی",
        "great": "پر-", "half": "سوتیلا-", "step": "سوتیلا-",
        "adoptive": "پرورش ", "foster": "ناظم ",
        "in_law": " داماد", "elder": "بڑا ", "younger": "چھوٹا ",
        "twin": "جڑواں ", "paternal": "پدری ", "maternal": "مادری ",
        "double": "دوہرا ", "ex": "سابقہ-",
        "god": "دینی-",
    },
    "es": {  # Spanish
        "name": "Spanish",
        "father": "padre", "mother": "madre", "parent": "progenitor",
        "son": "hijo", "daughter": "hija", "child": "hijo",
        "husband": "esposo", "wife": "esposa", "spouse": "cónyuge",
        "brother": "hermano", "sister": "hermana", "sibling": "hermano",
        "grandfather": "abuelo", "grandmother": "abuela",
        "grandson": "nieto", "granddaughter": "nieta",
        "uncle": "tío", "aunt": "tía",
        "nephew": "sobrino", "niece": "sobrina",
        "cousin": "primo",
        "great": "bis-", "half": "medio-", "step": "padrastro-",
        "adoptive": "adoptive ", "foster": "crianza ",
        "in_law": " político", "elder": "mayor ", "younger": "menor ",
        "twin": "gemelo ", "paternal": "paterno ", "maternal": "materno ",
        "double": "doble ", "ex": "ex-",
        "god": "padrino-",
    },
    "fr": {  # French
        "name": "French",
        "father": "père", "mother": "mère", "parent": "parent",
        "son": "fils", "daughter": "fille", "child": "enfant",
        "husband": "mari", "wife": "épouse", "spouse": "conjoint",
        "brother": "frère", "sister": "sœur", "sibling": "fratrie",
        "grandfather": "grand-père", "grandmother": "grand-mère",
        "grandson": "petit-fils", "granddaughter": "petite-fille",
        "uncle": "oncle", "aunt": "tante",
        "nephew": "neveu", "niece": "nièce",
        "cousin": "cousin",
        "great": "arrière-", "half": "demi-", "step": "beau-",
        "adoptive": "adoptive ", "foster": "d'accueil ",
        "in_law": " par alliance", "elder": "aîné ", "younger": "cadet ",
        "twin": "jumeau ", "paternal": "paternel ", "maternal": "maternel ",
        "double": "double ", "ex": "ex-",
        "god": "parrain-",
    },
    "de": {  # German
        "name": "German",
        "father": "Vater", "mother": "Mutter", "parent": "Elternteil",
        "son": "Sohn", "daughter": "Tochter", "child": "Kind",
        "husband": "Ehemann", "wife": "Ehefrau", "spouse": "Ehepartner",
        "brother": "Bruder", "sister": "Schwester", "sibling": "Geschwister",
        "grandfather": "Großvater", "grandmother": "Großmutter",
        "grandson": "Enkel", "granddaughter": "Enkelin",
        "uncle": "Onkel", "aunt": "Tante",
        "nephew": "Neffe", "niece": "Nichte",
        "cousin": "Cousin",
        "great": "Ur-", "half": "Stief-", "step": "Stief-",
        "adoptive": "Adoptiv-", "foster": "Pflege-",
        "in_law": " Schwager", "elder": "älterer ", "younger": "jüngerer ",
        "twin": "Zwillings-", "paternal": "väterlicher ", "maternal": "mütterlicher ",
        "double": "doppelter ", "ex": "Ex-",
        "god": "Paten-",
    },
    "ar": {  # Arabic
        "name": "Arabic",
        "father": "أب", "mother": "أم", "parent": "والد",
        "son": "ابن", "daughter": "ابنة", "child": "طفل",
        "husband": "زوج", "wife": "زوجة", "spouse": "قرين",
        "brother": "أخ", "sister": "أخت", "sibling": "شقيق",
        "grandfather": "جد", "grandmother": "جدة",
        "grandson": "حفيد", "granddaughter": "حفيدة",
        "uncle": "عم", "aunt": "عمة",
        "nephew": "ابن الأخ", "niece": "ابنة الأخ",
        "cousin": "ابن العم",
        "great": "جد-", "half": "نصف-", "step": "زوج الأب-",
        "adoptive": "بالتبني ", "foster": "حضانة ",
        "in_law": " المصاهرة", "elder": "أكبر ", "younger": "أصغر ",
        "twin": "توأم ", "paternal": "أبوي ", "maternal": "أموي ",
        "double": "مزدوج ", "ex": "سابق-",
        "god": "إله-",
    },
    "zh": {  # Chinese (Simplified)
        "name": "Chinese",
        "father": "父亲", "mother": "母亲", "parent": "父母",
        "son": "儿子", "daughter": "女儿", "child": "孩子",
        "husband": "丈夫", "wife": "妻子", "spouse": "配偶",
        "brother": "兄弟", "sister": "姐妹", "sibling": "兄弟姐妹",
        "grandfather": "祖父", "grandmother": "祖母",
        "grandson": "孙子", "granddaughter": "孙女",
        "uncle": "伯父", "aunt": "姑母",
        "nephew": "侄子", "niece": "侄女",
        "cousin": "堂兄弟",
        "great": "曾-", "half": "半-", "step": "继-",
        "adoptive": "养-", "foster": "寄养-",
        "in_law": " 姻亲", "elder": "大-", "younger": "小-",
        "twin": "双胞胎-", "paternal": "父系-", "maternal": "母系-",
        "double": "双重-", "ex": "前-",
        "god": "教-",
    },
    "ja": {  # Japanese
        "name": "Japanese",
        "father": "父", "mother": "母", "parent": "親",
        "son": "息子", "daughter": "娘", "child": "子供",
        "husband": "夫", "wife": "妻", "spouse": "配偶者",
        "brother": "兄弟", "sister": "姉妹", "sibling": "兄弟姉妹",
        "grandfather": "祖父", "grandmother": "祖母",
        "grandson": "孫", "granddaughter": "孫娘",
        "uncle": "伯父", "aunt": "伯母",
        "nephew": "甥", "niece": "姪",
        "cousin": "いとこ",
        "great": "曾-", "half": "異父異母-", "step": "継-",
        "adoptive": "養子-", "foster": "里親-",
        "in_law": " 婿", "elder": "年長-", "younger": "年少-",
        "twin": "双子-", "paternal": "父方-", "maternal": "母方-",
        "double": "二重-", "ex": "元-",
        "god": "名付け親-",
    },
    "ru": {  # Russian
        "name": "Russian",
        "father": "отец", "mother": "мать", "parent": "родитель",
        "son": "сын", "daughter": "дочь", "child": "ребёнок",
        "husband": "муж", "wife": "жена", "spouse": "супруг",
        "brother": "брат", "sister": "сестра", "sibling": "брат или сестра",
        "grandfather": "дедушка", "grandmother": "бабушка",
        "grandson": "внук", "granddaughter": "внучка",
        "uncle": "дядя", "aunt": "тётя",
        "nephew": "племянник", "niece": "племянница",
        "cousin": "двоюродный брат",
        "great": "пра-", "half": "сводн-", "step": "отчим-",
        "adoptive": "приёмн-", "foster": "приёмн-",
        "in_law": " зять", "elder": "старший ", "younger": "младший ",
        "twin": "близнец ", "paternal": "по отцу ", "maternal": "по матери ",
        "double": "двойной ", "ex": "бывш-",
        "god": "крёстн-",
    },
    "pt": {  # Portuguese
        "name": "Portuguese",
        "father": "pai", "mother": "mãe", "parent": "progenitor",
        "son": "filho", "daughter": "filha", "child": "criança",
        "husband": "marido", "wife": "esposa", "spouse": "cônjuge",
        "brother": "irmão", "sister": "irmã", "sibling": "irmão",
        "grandfather": "avô", "grandmother": "avó",
        "grandson": "neto", "granddaughter": "neta",
        "uncle": "tio", "aunt": "tia",
        "nephew": "sobrinho", "niece": "sobrinha",
        "cousin": "primo",
        "great": "bis-", "half": "meio-", "step": "padrasto-",
        "adoptive": "adotivo ", "foster": "acolhimento ",
        "in_law": " político", "elder": "mais velho ", "younger": "mais novo ",
        "twin": "gêmeo ", "paternal": "paterno ", "maternal": "materno ",
        "double": "duplo ", "ex": "ex-",
        "god": "padrinho-",
    },
    "id": {  # Indonesian
        "name": "Indonesian",
        "father": "ayah", "mother": "ibu", "parent": "orang tua",
        "son": "putra", "daughter": "putri", "child": "anak",
        "husband": "suami", "wife": "istri", "spouse": "pasangan",
        "brother": "saudara laki-laki", "sister": "saudara perempuan", "sibling": "saudara",
        "grandfather": "kakek", "grandmother": "nenek",
        "grandson": "cucu laki-laki", "granddaughter": "cucu perempuan",
        "uncle": "paman", "aunt": "bibi",
        "nephew": "keponakan laki-laki", "niece": "keponakan perempuan",
        "cousin": "sepupu",
        "great": "kakek-", "half": "tiri-", "step": "tiri-",
        "adoptive": "angkat ", "foster": "asuh ",
        "in_law": " ipar", "elder": "kakak ", "younger": "adik ",
        "twin": "kembar ", "paternal": "dari pihak ayah ", "maternal": "dari pihak ibu ",
        "double": "ganda ", "ex": "mantan-",
        "god": "wali-",
    },
    "vi": {  # Vietnamese
        "name": "Vietnamese",
        "father": "cha", "mother": "mẹ", "parent": "cha mẹ",
        "son": "con trai", "daughter": "con gái", "child": "con",
        "husband": "chồng", "wife": "vợ", "spouse": "vợ chồng",
        "brother": "anh em", "sister": "chị em", "sibling": "anh chị em",
        "grandfather": "ông nội", "grandmother": "bà nội",
        "grandson": "cháu trai", "granddaughter": "cháu gái",
        "uncle": "chú", "aunt": "cô",
        "nephew": "cháu trai", "niece": "cháu gái",
        "cousin": "anh em họ",
        "great": "cụ-", "half": "khác-", "step": "dượng-",
        "adoptive": "nuôi ", "foster": "nuôi ",
        "in_law": " ngoại", "elder": "anh ", "younger": "em ",
        "twin": "sinh đôi ", "paternal": "nội ", "maternal": "ngoại ",
        "double": "kép ", "ex": "cựu-",
        "god": "đạo-",
    },
    "tr": {  # Turkish
        "name": "Turkish",
        "father": "baba", "mother": "anne", "parent": "ebeveyn",
        "son": "oğul", "daughter": "kız", "child": "çocuk",
        "husband": "koca", "wife": "eş", "spouse": "eş",
        "brother": "erkek kardeş", "sister": "kız kardeş", "sibling": "kardeş",
        "grandfather": "dede", "grandmother": "babaanne",
        "grandson": "torun", "granddaughter": "torun",
        "uncle": "amca", "aunt": "hala",
        "nephew": "yeğen", "niece": "yeğen",
        "cousin": "kuzen",
        "great": "büyük-", "half": "üvey-", "step": "üvey-",
        "adoptive": "evlatlık ", "foster": "koruyucu ",
        "in_law": " damat", "elder": "büyük ", "younger": "küçük ",
        "twin": "ikiz ", "paternal": "baba tarafı ", "maternal": "anne tarafı ",
        "double": "çift ", "ex": "eski-",
        "god": "vaftiz-",
    },
    "ko": {  # Korean
        "name": "Korean",
        "father": "아버지", "mother": "어머니", "parent": "부모",
        "son": "아들", "daughter": "딸", "child": "자녀",
        "husband": "남편", "wife": "아내", "spouse": "배우자",
        "brother": "형제", "sister": "자매", "sibling": "형제자매",
        "grandfather": "할아버지", "grandmother": "할머니",
        "grandson": "손자", "granddaughter": "손녀",
        "uncle": "삼촌", "aunt": "고모",
        "nephew": "조카", "niece": "조카딸",
        "cousin": "사촌",
        "great": "증-", "half": "이부-", "step": "계-",
        "adoptive": "양자-", "foster": "위탁-",
        "in_law": " 사위", "elder": "큰-", "younger": "작은-",
        "twin": "쌍둥이-", "paternal": "친가-", "maternal": "외가-",
        "double": "이중-", "ex": "전-",
        "god": "대부-",
    },
    "it": {  # Italian
        "name": "Italian",
        "father": "padre", "mother": "madre", "parent": "genitore",
        "son": "figlio", "daughter": "figlia", "child": "figlio",
        "husband": "marito", "wife": "moglie", "spouse": "coniuge",
        "brother": "fratello", "sister": "sorella", "sibling": "fratello",
        "grandfather": "nonno", "grandmother": "nonna",
        "grandson": "nipote", "granddaughter": "nipote",
        "uncle": "zio", "aunt": "zia",
        "nephew": "nipote", "niece": "nipote",
        "cousin": "cugino",
        "great": "bis-", "half": "mezzo-", "step": "patrigno-",
        "adoptive": "adottivo ", "foster": "affidamento ",
        "in_law": " acquisito", "elder": "maggiore ", "younger": "minore ",
        "twin": "gemello ", "paternal": "paterno ", "maternal": "materno ",
        "double": "doppio ", "ex": "ex-",
        "god": "padrino-",
    },
}

# Post-process: add "late" (deceased) translation to every language.
# In most languages this is the same word used for "the late Mr. X".
_LATE_TRANSLATIONS = {
    "en": "late ", "hi": "स्वर्गीय ", "ta": "மறைந்த ", "te": "మరణించిన ",
    "bn": "স্বর্গীয় ", "mr": "दिवंगत ", "ml": "മരിച്ച ", "kn": "ಮೃತರಾದ ",
    "gu": "મૃત ", "pa": "ਮ੍ਰਿਤਕ ", "ur": "مرحوم ", "es": "difunto ",
    "fr": "feu ", "de": "verstorben ", "ar": "مرحوم ", "zh": "已故-",
    "ja": "故-", "ru": "покойн-", "pt": "falecido ",
    "id": "almarhum ", "vi": "cố-", "tr": "merhum ", "ko": "고-",
    "it": "fu ",
}
for _code, _late_word in _LATE_TRANSLATIONS.items():
    if _code in LANGUAGES:
        LANGUAGES[_code]["late"] = _late_word


# ---------------------------------------------------------------------------
# 4. COMPOSITION RULES — translate English seed term → language term
# ---------------------------------------------------------------------------

def compose_term(lang_dict: Dict[str, str], english_term: str, sig: KinshipSignature) -> str:
    """Rule-based composition of the localized kinship term.

    Strategy:
      1. Decompose the English term into modifier prefix + root + suffix.
      2. Map each piece via lang_dict.
      3. Reassemble using language-specific glue.
    """
    if lang_dict.get("name") == "English":
        return english_term

    parts = english_term.lower().split()
    if not parts:
        return english_term

    # Detect consanguinity/seniority markers in english phrase
    prefix_parts = []
    suffix_parts = []
    root_parts = []

    skip_words = {"cousin", "(m)", "(f)", "[cousin", "(m)]", "(f)]", "[m]", "[f]",
                  "sibling's", "spouse's", "own", "child's", "grandchild's"}
    i = 0
    while i < len(parts):
        w = parts[i]
        if w in {"elder", "younger", "twin"}:
            prefix_parts.append(lang_dict.get(w, w + " "))
        elif w in {"half", "step", "adoptive", "foster", "double", "ex", "god", "late"}:
            prefix_parts.append(lang_dict.get(w, w + "-"))
        elif w in {"paternal", "maternal"}:
            prefix_parts.append(lang_dict.get(w, w + " "))
        elif w in {"great-", "2nd", "3rd", "4th", "5th", "6th", "7th"}:
            # Treat as great-marker
            prefix_parts.append(lang_dict.get("great", "great-"))
        elif w in {"once", "twice", "thrice"}:
            # removal marker — append as suffix
            suffix_parts.append(w + " removed")
        elif w in {"removed"}:
            pass
        elif w in skip_words:
            pass  # ignore
        else:
            # This is the root word; map to language
            # Handle compound: "father-in-law", "step-grandfather", etc.
            w_clean = w.rstrip(",;")
            root_parts.append(_map_root(w_clean, lang_dict))
        i += 1

    if not root_parts:
        return english_term  # fallback

    # Compose: prefix + root + suffix + in_law marker
    in_law_marker = ""
    if "in-law" in english_term.lower() or "in_law" in english_term.lower():
        in_law_marker = lang_dict.get("in_law", "-in-law")
        # Strip "-in-law" from roots if present
        root_parts = [r.replace("-in-law", "").replace("_in_law", "") for r in root_parts]

    prefix_str = "".join(prefix_parts)
    root_str = " ".join(root_parts)
    suffix_str = " ".join(suffix_parts)

    # For in-law, language-specific glue
    if in_law_marker:
        # e.g. Japanese: 息子の婿 (root + の婿), Hindi: दामाद (single word often)
        if lang_dict.get("name") in {"Hindi", "Tamil", "Telugu", "Bengali", "Marathi",
                                      "Malayalam", "Kannada", "Gujarati", "Punjabi", "Urdu"}:
            # Indic languages often have unique in-law words; use marker as suffix
            composed = f"{prefix_str}{root_str}{in_law_marker}".strip()
        else:
            composed = f"{prefix_str}{root_str}{in_law_marker}".strip()
    else:
        composed = f"{prefix_str}{root_str}".strip()

    if suffix_str:
        composed = f"{composed} {suffix_str}".strip()

    return composed.strip()


def _map_root(w: str, lang_dict: Dict[str, str]) -> str:
    """Map an English root to its language equivalent, handling hyphenated forms."""
    # Direct lookup
    direct_map = {
        "father": "father", "mother": "mother", "parent": "parent",
        "son": "son", "daughter": "daughter", "child": "child",
        "husband": "husband", "wife": "wife", "spouse": "spouse",
        "brother": "brother", "sister": "sister", "sibling": "sibling",
        "grandfather": "grandfather", "grandmother": "grandmother",
        "grandson": "grandson", "granddaughter": "granddaughter",
        "uncle": "uncle", "aunt": "aunt",
        "nephew": "nephew", "niece": "niece",
        "cousin": "cousin",
    }
    if w in direct_map:
        return lang_dict.get(direct_map[w], w)
    # Hyphenated: prefix + root
    if "-" in w:
        pieces = w.split("-")
        if len(pieces) == 2:
            modifier, root = pieces
            mod_key = modifier if modifier in lang_dict else "great"
            root_val = _map_root(root, lang_dict)
            return f"{lang_dict.get(mod_key, modifier + '-')}{root_val}"
    # Compound like "grandfather"
    for prefix in ("grand", "great"):
        if w.startswith(prefix) and w != prefix:
            root_key = w[len(prefix):]
            if root_key in direct_map:
                return f"{lang_dict.get('great', prefix)}{lang_dict.get(direct_map[root_key], root_key)}"
    return lang_dict.get(w, w)


# ---------------------------------------------------------------------------
# 5. ROW ASSEMBLY
# ---------------------------------------------------------------------------

@dataclass
class VocabRow:
    row_id: int
    canonical_id: str
    signature_key: str
    path_pattern: str
    generation_delta: int
    side: str
    consanguinity: str
    gender_anchor: str
    seniority: str
    removal: int
    double_kinship: bool
    temporal: str           # current | former | late  (NEW)
    category: str
    english_term: str
    notes: str
    language_code: str
    language_name: str
    localized_term: str
    variant_type: str       # primary | regional | dialectal
    variant_rank: int       # 0 = primary (engine default), 1+ = alternates


def build_rows() -> List[VocabRow]:
    seeds = _build_seed_concepts()
    rows: List[VocabRow] = []
    row_id = 1
    for canonical_id, english_term, sig, category, notes, variant_type in seeds:
        for lang_code, lang_dict in LANGUAGES.items():
            localized = compose_term(lang_dict, english_term, sig)
            rank = 0 if variant_type == "primary" else (1 if variant_type == "regional" else 2)
            rows.append(VocabRow(
                row_id=row_id,
                canonical_id=canonical_id,
                signature_key=sig.signature_key(),
                path_pattern=sig.pathPattern,
                generation_delta=sig.generationDelta,
                side=sig.side,
                consanguinity=sig.consanguinity,
                gender_anchor=sig.genderAnchor,
                seniority=sig.seniority,
                removal=sig.removal,
                double_kinship=sig.doubleKinship,
                temporal=sig.temporal,
                category=category,
                english_term=english_term,
                notes=notes,
                language_code=lang_code,
                language_name=lang_dict["name"],
                localized_term=localized,
                variant_type=variant_type,
                variant_rank=rank,
            ))
            row_id += 1
    return rows


# ---------------------------------------------------------------------------
# 6. OUTPUT WRITERS
# ---------------------------------------------------------------------------

OUTPUT_DIR = Path("/home/z/my-project/download")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def write_xlsx(rows: List[VocabRow]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Kinship Vocabulary"

    headers = [
        "row_id", "canonical_id", "signature_key", "path_pattern",
        "generation_delta", "side", "consanguinity", "gender_anchor",
        "seniority", "removal", "double_kinship", "temporal",
        "category", "english_term", "notes",
        "language_code", "language_name", "localized_term",
        "variant_type", "variant_rank",
    ]
    ws.append(headers)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Body
    for r in rows:
        ws.append([
            r.row_id, r.canonical_id, r.signature_key, r.path_pattern,
            r.generation_delta, r.side, r.consanguinity, r.gender_anchor,
            r.seniority, r.removal, r.double_kinship, r.temporal,
            r.category, r.english_term, r.notes,
            r.language_code, r.language_name, r.localized_term,
            r.variant_type, r.variant_rank,
        ])

    # Auto column widths (cap at 50)
    for col_idx, header in enumerate(headers, start=1):
        max_len = max(
            len(str(header)),
            max((len(str(getattr(r, header) if hasattr(r, header) else "")) for r in rows), default=0)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Freeze header
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Second sheet: signature summary
    summary_ws = wb.create_sheet("Signature Summary")
    summary_ws.append(["category", "count", "languages"])
    for cat in sorted({r.category for r in rows}):
        cat_rows = [r for r in rows if r.category == cat]
        summary_ws.append([
            cat,
            len(cat_rows),
            ", ".join(sorted({r.language_name for r in cat_rows})),
        ])
    for col_idx in range(1, 4):
        cell = summary_ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    summary_ws.column_dimensions["A"].width = 24
    summary_ws.column_dimensions["B"].width = 10
    summary_ws.column_dimensions["C"].width = 80

    # Third sheet: language coverage
    lang_ws = wb.create_sheet("Language Coverage")
    lang_ws.append(["language_code", "language_name", "row_count"])
    for code, d in LANGUAGES.items():
        lang_ws.append([code, d["name"], sum(1 for r in rows if r.language_code == code)])
    for col_idx in range(1, 4):
        cell = lang_ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    lang_ws.column_dimensions["A"].width = 16
    lang_ws.column_dimensions["B"].width = 24
    lang_ws.column_dimensions["C"].width = 14

    out_path = OUTPUT_DIR / "daxelo_kinrel_vocabulary.xlsx"
    wb.save(out_path)
    return out_path


def write_json(rows: List[VocabRow]) -> Path:
    out_path = OUTPUT_DIR / "daxelo_kinrel_vocabulary.json"
    payload = {
        "engine": "Daxelo-Kinrel Deterministic Kinship Engine v3.0",
        "module": "Vocabulary Mapper",
        "version": "3.0",
        "total_terms": len(rows),
        "primary_terms_only": sum(1 for r in rows if r.variant_rank == 0),
        "languages": [{"code": k, "name": v["name"]} for k, v in LANGUAGES.items()],
        "schema": {
            "row_id": "integer, 1-indexed",
            "canonical_id": "string — PARENT | SPOUSE | ADOPTIVE_PARENT | STEP_PARENT | DERIVED",
            "signature_key": "string — composite signature for deterministic lookup",
            "path_pattern": "string — traversal primitives joined by _",
            "generation_delta": "integer -8..+8 (negative = ancestor, positive = descendant)",
            "side": "paternal | maternal | none",
            "consanguinity": "blood | half | step | adoptive | inLaw | foster | spiritual",
            "gender_anchor": "male | female | neutral",
            "seniority": "elder | younger | twin | none",
            "removal": "integer 0..8 (for cousin removals)",
            "double_kinship": "boolean — true for double cousins etc.",
            "temporal": "current | former | late — first-class temporal state (NEW in v3.1)",
            "category": "string — classification of the relationship",
            "english_term": "string — reference English term",
            "notes": "string — human-readable description",
            "language_code": "ISO 639-1 code",
            "language_name": "Display name",
            "localized_term": "string — the kinship term in the target language",
            "variant_type": "primary | regional | dialectal — disambiguates synonyms",
            "variant_rank": "integer — 0 = primary (engine default), 1+ = regional/dialectal alternates",
        },
        "lookup_contract": (
            "For deterministic lookup, the engine MUST filter by variant_rank=0 first. "
            "Only fall back to variant_rank>=1 if the user explicitly requests a regional/dialectal variant."
        ),
        "rows": [asdict(r) for r in rows],
    }
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def write_csv(rows: List[VocabRow]) -> Path:
    out_path = OUTPUT_DIR / "daxelo_kinrel_vocabulary.csv"
    headers = [
        "row_id", "canonical_id", "signature_key", "path_pattern",
        "generation_delta", "side", "consanguinity", "gender_anchor",
        "seniority", "removal", "double_kinship", "temporal",
        "category", "english_term", "notes",
        "language_code", "language_name", "localized_term",
        "variant_type", "variant_rank",
    ]
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in rows:
            writer.writerow([
                r.row_id, r.canonical_id, r.signature_key, r.path_pattern,
                r.generation_delta, r.side, r.consanguinity, r.gender_anchor,
                r.seniority, r.removal, r.double_kinship, r.temporal,
                r.category, r.english_term, r.notes,
                r.language_code, r.language_name, r.localized_term,
                r.variant_type, r.variant_rank,
            ])
    return out_path


# ---------------------------------------------------------------------------
# 7. MAIN
# ---------------------------------------------------------------------------

def main():
    print("[1/4] Building seed concepts...")
    seeds = _build_seed_concepts()
    print(f"      seed concepts (English): {len(seeds)}")

    print("[2/4] Building vocabulary rows (seed x language matrix)...")
    rows = build_rows()
    print(f"      total vocabulary rows: {len(rows)}")

    print("[3/4] Writing XLSX / JSON / CSV...")
    xlsx_path = write_xlsx(rows)
    json_path = write_json(rows)
    csv_path = write_csv(rows)

    print("[4/4] Done.")
    print(f"  XLSX: {xlsx_path}  ({xlsx_path.stat().st_size:,} bytes)")
    print(f"  JSON: {json_path}  ({json_path.stat().st_size:,} bytes)")
    print(f"  CSV : {csv_path}   ({csv_path.stat().st_size:,} bytes)")

    # Determinism check — primary terms only (rank 0)
    print("\nDeterminism check (primary terms only, variant_rank=0):")
    primary_rows = [r for r in rows if r.variant_rank == 0]
    sig_to_terms: Dict[str, set] = {}
    for r in primary_rows:
        key = (r.signature_key, r.language_code)
        sig_to_terms.setdefault(key, set()).add(r.localized_term)
    ambiguous = {k: v for k, v in sig_to_terms.items() if len(v) > 1}
    print(f"  ambiguous (signature, language) pairs: {len(ambiguous)}")
    if ambiguous:
        for k, v in list(ambiguous.items())[:5]:
            print(f"    {k} -> {v}")

    # Coverage by category
    print("\nCategory breakdown:")
    cat_counts: Dict[str, int] = {}
    for r in rows:
        cat_counts[r.category] = cat_counts.get(r.category, 0) + 1
    for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
        print(f"  {cat:24s} {n:>6d}")

    # Language coverage
    print("\nLanguage coverage:")
    lang_counts: Dict[str, int] = {}
    for r in rows:
        lang_counts[r.language_name] = lang_counts.get(r.language_name, 0) + 1
    for lang, n in sorted(lang_counts.items()):
        print(f"  {lang:14s} {n:>6d}")

    # Sanity: meet target?
    target = 5396
    if len(rows) >= target:
        print(f"\n✓ TARGET MET: {len(rows)} rows >= {target} target")
    else:
        print(f"\n✗ TARGET MISSED: {len(rows)} rows < {target} target — expand seed or languages")


if __name__ == "__main__":
    main()
